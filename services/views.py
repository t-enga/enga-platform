from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from services.forms import RouteForm,RoutePointForm, CreateTripForm, TravelDataForm, TraficDataForm, TravelExpencesForm, BillinForm, RanchExit, ArriveSupplier,ArrivePoint, ArriveRanch, ExitSupplier, Cargo1, Cargo2, Cargo3 
from services.models import Trips, IdConcept, MonthExpenses, Month,RoutePoint, Route
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
import io
import json
import requests
from services.decorators import group_required
from datetime import datetime, date
from django.utils import timezone

#views
def home(request):
    return render(request, 'home.html')

def my_view(request):
    user = request.user
    is_administrative = user.groups.filter(name='Administrative').exists()

    return render(request, 'my_template.html', {'is_administrative': is_administrative})

@group_required('Administrative', 'SuperUsers')
@login_required  # Esto asegura que el usuario ya ha iniciado sesión antes de acceder a la vista
def principal(request):
    # Aquí verifica los grupos del usuario y redirige en consecuencia
    if request.user.groups.filter(name='Administrative').exists():
        return redirect('/')
    elif request.user.groups.filter(name='SuperUsers').exists():
        return redirect('/')
    elif request.user.groups.filter(name='Drivers').exists():
        return redirect('status_supplie')
    elif request.user.groups.filter(name='Supplies').exists():
        return redirect('status_table')
    else:
        # Si el usuario no pertenece a ningún grupo específico, se redirigirige a una página genérica o mostrar un mensaje de error.
        return render(request, 'home.html', {'error_message': 'Usuario no tiene un grupo válido'})

@group_required('Administrative', 'SuperUsers')
@login_required
def map(request):
    return render(request, 'map.html')

def signin(request):
    if request.method == 'GET':
        form = AuthenticationForm()
        return render(request, 'signin.html', {'form': form})
    else:
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = authenticate(request, username=form.cleaned_data['username'], password=form.cleaned_data['password'])
            if user is not None:
                login(request, user)
                return redirect('principal')  # Redirige a la vista principal
            else:
                return render(request, 'signin.html', {
                    'form': form,
                    'error': 'Usuario o contraseña incorrecto'
                })
        else:
            return render(request, 'signin.html', {
                'form': form
            })

@group_required('Administrative', 'SuperUsers')
@login_required
def export_to_excel(request):
    data = Trips.objects.all()
    wb = Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
    
    field_labels = {
        'lote': 'Número de Lote',
        'week': 'Semana',
        'weight': 'Peso',
        'type_service': 'Tipo de Servicio',
        'month': 'Mes',
        'serv_date': 'Fecha de Servicio',
        'supplier': 'Proveedor',
        'oigin': 'Origen',
        'meeting_place': 'Lugar de Encuentro',
        'meeting_hour': 'Hora de Encuentro',
        'truck': 'Camión',
        'truck_plates': 'Placas de Camión',
        'trailer': 'Remolque',
        'trailer_plates': 'Placas de Remolque',
        'driver': 'Conductor',
        'ranch_exit': 'Salida de Rancho',
        'kms_exit': 'Kilómetros de Salida',
        'arrive_point': 'Punto de Llegada',
        'arrive_kms': 'Kilómetros de Llegada',
        'cargo1_arrival': 'Llegada de Carga 1',
        'cargo1_kms': 'Kilómetros de Carga 1',
        'cargo1_site': 'Sitio de Carga 1',
        'cargo1_collected': 'Carga 1 Recogida',
        'cargo1_fin': 'Finalización de Carga 1',
        'cargo2_arrival': 'Llegada de Carga 2',
        'cargo2_kms': 'Kilómetros de Carga 2',
        'cargo2_site': 'Sitio de Carga 2',
        'cargo2_collected': 'Carga 2 Recogida',
        'cargo2_fin': 'Finalización de Carga 2',
        'cargo3_arrival': 'Llegada de Carga 3',
        'cargo3_kms': 'Kilómetros de Carga 3',
        'cargo3_site': 'Sitio de Carga 3',
        'cargo3_collected': 'Carga 3 Recogida',
        'cargo3_fin': 'Finalización de Carga 3',
        'exit_supplier': 'Salida de Proveedor',
        'comentaries': 'Comentarios',
        'ranch_arrive': 'Llegada a Rancho',
        'ranch_kms': 'Kilómetros de Llegada a Rancho',
        'total_collected': 'Total Recogido',
        'total_kms': 'Kilómetros Totales',
        'empty_kms': 'Kilómetros Vacíos',
        'dead_kms': 'Kilómetros Muertos',
        'lts_diesel': 'Litros de Diesel',
        'performance': 'Rendimiento',
        'urea': 'Urea',
        'stand': 'Soporte',
        'iave': 'IAVE',
        'food': 'Comida',
        'dessinfections': 'Desinfecciones',
        'fixes': 'Reparaciones',
        'parts': 'Piezas',
        'tires': 'Llantas',
        'pension': 'Pensión',
        'federal': 'Federal',
        'type_fact': 'Tipo de Factura',
        'fact_cp': 'Código Postal de la Factura',
        'date_fact': 'Fecha de Factura',
        'amount': 'Monto',
        'unit_price': 'Precio Unitario',
        'type_note': 'Tipo de Nota',
        'amount_note': 'Monto de Nota',
    }

    for col_num, (field_name, label) in enumerate(field_labels.items(), start=1):
        ws.cell(row=1, column=col_num, value=label)

    for row_num, trip in enumerate(data, start=2):
        for col_num, field_name in enumerate(field_labels.keys(), start=1):
            field_value = getattr(trip, field_name)
            if field_name == 'oigin'and field_value is not None:
                field_value = field_value.name_origin
            elif field_name in ['type_service', 'month', 'truck', 'truck_plates', 'trailer', 'trailer_plates', 'driver', 'supplier']:
                field_value = str(field_value)
            elif field_name == 'serv_date':
                if field_value is not None:
                    if not isinstance(field_value, date):
                        field_value = field_value.date()  # Convierte a datetime.date si no lo es
                    field_value = datetime.combine(field_value, datetime.min.time()).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    field_value = ""
            elif isinstance(field_value, datetime):
                # Convierte las fechas y horas a texto sin formato de zona horaria
                field_value = field_value.strftime('%Y-%m-%d %H:%M:%S')

            ws.cell(row=row_num, column=col_num, value=field_value)
        
    output = io.BytesIO()
    wb.save(output)
    response.write(output.getvalue())
    return response

@login_required
@group_required('Administrative', 'SuperUsers')
def signup(request):
    form = UserCreationForm
    if request.method == 'GET':
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
           try:
                user = User.objects.create_user(username=request.POST['username'], password=request.POST['password1'])
                user.save()
                return render(request, 'signup.html', {
                    'form': UserCreationForm
                })
           except IntegrityError:
                return render(request, 'signup.html', {
                    'error': 'Usuario ya existe',
                    'form': UserCreationForm
                })
        else:
            return render(request, 'signup.html', {
                    'error': 'Contraseñas no coinciden',
                    'form': UserCreationForm
                })

@login_required
@group_required('Administrative', 'SuperUsers')
def trips(request):
    trips = Trips.objects.filter(ranch_arrive__isnull=True)
    return render(request, 'trips.html',{
        'trips': trips
    })

@login_required 
def trips_drivers(request):
    current_user = request.user.username
    trips = Trips.objects.filter(
        ranch_arrive__isnull=True,
        truck__truck_num=current_user
    )

    return render(request, 'trips_driver.html', {
        'trips': trips
    })

#Datos de viajes(horarios)
@login_required
def drivers_trip(request, id_service):
    trip = get_object_or_404(Trips, pk=id_service)

    form_ranch_exit = RanchExit(instance=trip)
    form_arrive_point = ArrivePoint(instance=trip)
    form_arrive_supplier = ArriveSupplier(instance=trip)
    form_exit_supplier = ExitSupplier(instance=trip)
    form_arrive_ranch = ArriveRanch(instance=trip)

    if request.method == 'POST':
        if 'submit_ranch_exit' in request.POST:
            form_ranch_exit = RanchExit(request.POST, instance=trip)
            if form_ranch_exit.is_valid():
                form_ranch_exit.instance.ranch_exit = timezone.now()
                form_ranch_exit.save()
                return redirect('trips_drivers')

        elif 'submit_arrive_point' in request.POST:
            form_arrive_point = ArrivePoint(request.POST, instance=trip)
            if form_arrive_point.is_valid():
                form_arrive_point.instance.arrive_point = timezone.now()
                form_arrive_point.save()
                return redirect('trips_drivers')

        elif 'submit_arrive_supplier' in request.POST:
            form_arrive_supplier = ArriveSupplier(request.POST, instance=trip)
            if form_arrive_supplier.is_valid():
                form_arrive_supplier.instance.arrive_supp = timezone.now()
                form_arrive_supplier.save()
                return redirect('trips_drivers')
            
        elif 'submit_exit_supplier' in request.POST:
            form_exit_supplier = ExitSupplier(request.POST, instance=trip)
            if form_exit_supplier.is_valid():
                form_exit_supplier.instance.exit_supplier = timezone.now()
                form_exit_supplier.save()
                return redirect('trips_drivers')
            
        elif 'submit_arrive_ranch' in request.POST:
            form_arrive_ranch = ArriveRanch(request.POST, instance=trip)
            if form_arrive_ranch.is_valid():
                form_arrive_ranch.instance.ranch_arrive = timezone.now()
                form_arrive_ranch.save()
                return redirect('trips_drivers')
            
    return render(request, 'drivers_trip.html', {
        'trip': trip,
        'form_ranch_exit': form_ranch_exit,
        'form_arrive_point': form_arrive_point,
        'form_arrive_supplier': form_arrive_supplier,
        'form_exit_supplier': form_exit_supplier,
        'form_arrive_ranch': form_arrive_ranch,
    })

@login_required
def drivers_cargo(request, id_service):
    trip = get_object_or_404(Trips, pk=id_service)
    form_cargo1 = Cargo1(instance=trip)
    form_cargo2 = Cargo2(instance=trip)
    form_cargo3 = Cargo3(instance=trip)
    
    if request.method == 'POST':
        if 'submit_cargo1' in request.POST:
            form_cargo1 = Cargo1(request.POST, instance=trip)
            if form_cargo1.is_valid():
                cargo1_kms = form_cargo1.cleaned_data.get('cargo1_kms')
                cargo1_collected = form_cargo1.cleaned_data.get('cargo1_collected')
                if trip.cargo1_arrival == None and cargo1_kms is not None:
                    trip.cargo1_arrival = timezone.now()
                    
                if trip.cargo1_fin == None and cargo1_collected is not None:
                    trip.cargo1_fin = timezone.now()
                
                trip.save()
                form_cargo1.save()
                return redirect('drivers_cargo', id_service=trip.pk)
        
        elif 'submit_cargo2' in request.POST:
            form_cargo2 = Cargo2(request.POST, instance=trip)
            if form_cargo2.is_valid():
                cargo2_kms = form_cargo2.cleaned_data.get('cargo2_kms')
                cargo2_collected = form_cargo2.cleaned_data.get('cargo2_collected')
                if trip.cargo2_arrival == None and cargo2_kms is not None:
                    trip.cargo2_arrival = timezone.now()
                    
                if trip.cargo2_fin == None and cargo2_collected is not None:
                    trip.cargo2_fin = timezone.now()
                
                trip.save()
                form_cargo2.save()
                return redirect('drivers_cargo', id_service=trip.pk)
            
        elif 'submit_cargo3' in request.POST:
            form_cargo3 = Cargo3(request.POST, instance=trip)
            if form_cargo3.is_valid():
                cargo3_kms = form_cargo3.cleaned_data.get('cargo3_kms')
                cargo3_collected = form_cargo3.cleaned_data.get('cargo3_collected')
                if trip.cargo3_arrival == None and cargo3_kms is not None:
                    trip.cargo3_arrival = timezone.now()
                    
                if trip.cargo3_fin == None and cargo3_collected is not None:
                    trip.cargo3_fin = timezone.now()
                
                trip.save()
                form_cargo3.save()
                return redirect('drivers_cargo', id_service=trip.pk)
            
    return render(request, 'trip_cargo.html', {
        'trip': trip,
        'form_cargo1': form_cargo1,
        'form_cargo2': form_cargo2,
        'form_cargo3': form_cargo3,
    })

#<<<<------------FIN------------->>>>>>

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_create(request):
    routes = Route.objects.all()

    if request.method == 'POST':
        form = CreateTripForm(request.POST)
        if form.is_valid():
            new_trip = form.save(commit=False)
            
            # Asignar la ruta seleccionada al nuevo viaje
            route_id = request.POST.get('route')  # Obtener el ID de la ruta desde el formulario
            route = Route.objects.get(id_route=route_id)  # Obtener la instancia de la ruta
            new_trip.route = route  # Asignar la ruta al nuevo viaje

            new_trip.save()
            return redirect('trips')
    else:
        form = CreateTripForm()

    return render(request, 'trip_create.html', {
        'form': form,
        'routes': routes
    })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_travel(request):
    trips = Trips.objects.filter(ranch_arrive__isnull=True)
    return render(request, 'trip_travel.html',{
        'trips': trips
    })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_finished(request):
    trips = Trips.objects.filter(ranch_arrive__isnull=False).order_by('lote')
    return render(request, 'trip_finished.html',{
        'trips': trips
    })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_edit(request, id_service):
    if request.method == 'GET':
        trips = get_object_or_404(Trips, pk=id_service)
        form = CreateTripForm(instance=trips)
        return render(request, 'trip_edit.html', {
        'trip': trips,
        'form': form
        })
    else:
        try:
            trips = get_object_or_404(Trips, pk=id_service)
            form = CreateTripForm(request.POST, instance=trips)
            form.save()
            return redirect('trips')
        except ValueError:
            return render(request, 'trips.html', {
            'trip': trips,   
            'form': form,
            'error': 'Error actualizando tarea'
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_billing(request, id_service):
    if request.method == 'GET':
        trips = get_object_or_404(Trips, pk=id_service)
        form = BillinForm(instance=trips)
        return render(request, 'trip_billing.html', {
        'trip': trips,
        'form': form
        })
    else:
        try:
            trips = get_object_or_404(Trips, pk=id_service)
            form = BillinForm(request.POST, instance=trips)
            form.save()
            return redirect('trips')
        except ValueError:
            return render(request, 'trip_billing.html', {
            'trip': trips,   
            'form': form,
            'error': 'Error al ingresar datos'
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_trafic(request, id_service):
    if request.method == 'GET':
        trips = get_object_or_404(Trips, pk=id_service)
        form = TraficDataForm(instance=trips)
        return render(request, 'trip_trafic.html', {
        'trip': trips,
        'form': form
        })
    else:
        try:
            trips = get_object_or_404(Trips, pk=id_service)
            form = TraficDataForm(request.POST, instance=trips)
            form.save()
            return redirect('trips')
        except ValueError:
            return render(request, 'trip_trafic.html', {
            'trip': trips,   
            'form': form,
            'error': 'Error actualizando tarea'
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_expences(request, id_service):
    if request.method == 'GET':
        trips = get_object_or_404(Trips, pk=id_service)
        form = TravelExpencesForm(instance=trips)
        return render(request, 'trip_expences.html', {
        'trip': trips,
        'form': form
        })
    else:
        try:
            trips = get_object_or_404(Trips, pk=id_service)
            form = TravelExpencesForm(request.POST, instance=trips)
            form.save()
            return redirect('trips')
        except ValueError:
            return render(request, 'trip_expences.html', {
            'trip': trips,   
            'form': form,
            'error': 'Error actualizando tarea'
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_detail(request, id_service):
    if request.method == 'GET':
        trips = get_object_or_404(Trips, pk=id_service)
        form = TravelDataForm(instance=trips)
        return render(request, 'trip_detail.html', {
        'trip': trips,
        'form': form
        })
    else:
        try:
            trips = get_object_or_404(Trips, pk=id_service)
            form = TravelDataForm(request.POST, instance=trips)
            form.save()
            return redirect('trips')
        except ValueError:
            return render(request, 'trip_detail.html', {
            'trip': trips,   
            'form': form,
            'error': 'Error actualizando tarea'
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def trip_delete(request, id_service):
    trip = get_object_or_404(Trips, pk=id_service)
    if request.method == 'POST':
        trip.delete()
        return redirect(trips)

@login_required
def signout(request):
    logout(request)
    return redirect('home')

@login_required
@group_required('Administrative', 'SuperUsers')
def get_coordinates(request):
    if request.method == 'GET':
        client_id = 2428
        user_id = 53047
        api_key = "AIzaSyAxI0HC4Wf8J2iuTBE1UuznBRgII5E2Tl4"  

        url = f"https://telemetry.dev.api.enlacefl.com/clients/{client_id}/users/{user_id}/assets/current-position?key={api_key}"
        
        response = requests.get(url)
        data = response.json()

        coordinates = []
        if 'data' in data:
            coordinates = []
            for device in data['data']:
                latitude = device['position']['latitude']
                longitude = device['position']['longitude']
                device_name = device.get('vehicleNumber', 'numberPlates')
                coordinates.append({'latitude': latitude, 'longitude': longitude, 'name': device_name})
                
            return JsonResponse({'coordinates': coordinates})
        else:
            return JsonResponse({'error': 'No se encontraron datos válidos'})

def user_belongs_to_group(user, group_name):
    return user.groups.filter(name=group_name).exists()

@login_required
def your_view(request):
    return render(request, 'base.html', {
        'user_belongs_to_group': user_belongs_to_group,
    })

@login_required
def trip_status(request):
    return render(request,'trip_status.html')

@login_required
def expenses_table(request):
    concepts = IdConcept.objects.all()
    months = Month.objects.all()
    years = [2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040]

    # Inicialmente, muestra una tabla vacía
    expenses = []

    return render(request, 'expenses_table.html', {'expenses': expenses, 'concepts': concepts, 'months': months, 'years': years})

@login_required
def filter_expenses(request):
    year = request.GET.get('year')
    month = request.GET.get('month')

    # Si ambos year y month tienen valores, filtra por ambos. 
    # Si solo uno tiene valor, filtra solo por ese.
    if year and month:
        expenses = MonthExpenses.objects.filter(year=year, month=month)
    elif year:
        expenses = MonthExpenses.objects.filter(year=year)
    elif month:
        expenses = MonthExpenses.objects.filter(month=month)
    else:
        # Si ninguno tiene valor, retorna un JsonResponse vacío.
        return JsonResponse([], safe=False)

    data = []
    for expense in expenses:
        data.append({
            'concept': expense.name_conceptt.name_concept,  # Asegúrate de acceder al nombre del concepto
            'month': expense.month.id_month,
            'year': expense.year,
            'total': expense.total,
            'expense_month': expense.expense_month,
            'proportional_trip': expense.proportional_trip,
            'proportional_km': expense.proportional_km
        })

    return JsonResponse(data, safe=False)

def table_ganade(request):
    trips = Trips.objects.all()
    return render(request, 'table_ganade.html', {'trips': trips})

def board_cig(request):
    return render(request, 'board_mci.html')

def get_info_gps(request):
    if request.method == 'GET':
        client_id = 2428
        user_id = 53047
        api_key = "AIzaSyAxI0HC4Wf8J2iuTBE1UuznBRgII5E2Tl4"  

        url = f"https://telemetry.dev.api.enlacefl.com/clients/{client_id}/users/{user_id}/assets/current-position?key={api_key}"
        
        response = requests.get(url)
        data = response.json()

        if 'data' in data:
            return render(request, 'get_info_gps.html', {'data': data['data']})
        else:
            return render(request, 'get_info_gps.html', {'error': "Datos no encontrados"})

def create_route(request):
    if request.method == 'POST':
        route_form = RouteForm(request.POST)
        if route_form.is_valid():
            route = route_form.save()
            return JsonResponse({'message': 'Ruta guardada correctamente'})
        else:
            return JsonResponse({'error': 'El formulario no es válido.'}, status=400)
    else:
        route_form = RouteForm()

    return render(request, 'create_route.html', {'route_form': route_form})

def save_route(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nombre = data.get('nombre')
            descripcion = data.get('descripcion')
            puntos = data.get('puntos')
            
            if not nombre or not descripcion:
                return JsonResponse({'error': 'Debes ingresar un nombre y una descripción para la ruta.'}, status=400)

            if len(puntos) < 2:
                return JsonResponse({'error': 'Debes seleccionar al menos dos puntos para formar una ruta.'}, status=400)

            # Crea la nueva ruta
            ruta = Route.objects.create(name_route=nombre, desc_route=descripcion)

            # Guarda los puntos asociados a la ruta
            for punto in puntos:
                latitud = punto['lat']
                longitud = punto['lng']
                RoutePoint.objects.create(route=ruta, lat_route=latitud, lon_route=longitud)

            # Enviar una respuesta
            return JsonResponse({'message': 'Ruta y puntos guardados correctamente'})
        except json.decoder.JSONDecodeError as e:
            return JsonResponse({'error': 'Error al decodificar JSON.'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Método no permitido'}, status=405)

def routes_assign(request, id_service):
    #client_id = 2428
    #user_id = 53047
    #api_key = "AIzaSyAxI0HC4Wf8J2iuTBE1UuznBRgII5E2Tl4"
    #url = f"https://telemetry.dev.api.enlacefl.com/clients/{client_id}/users/{user_id}/assets/current-position?key={api_key}"
    #response = requests.get(url)
    #data = response.json()
    
    trip = get_object_or_404(Trips, id_service=id_service)
    route_points = RoutePoint.objects.filter(route_id=trip.route_id)
    route_coordinates = [{'lat': point.lat_route, 'lng': point.lon_route} for point in route_points]

    context = {
        'trip': trip,
        'route_coordinates': route_coordinates,
        #'data': data['data'],
    }

    return render(request, 'routes_assign.html', context)
    
def get_route_points(request, id_route):
    route = get_object_or_404(Route, id_route=id_route)
    points = list(route.routepoint_set.values('lat_route', 'lon_route'))
    return JsonResponse(points, safe=False)

 