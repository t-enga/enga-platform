from django.shortcuts import render, redirect, get_object_or_404
from .forms import CreateSupplieForm, StatusForm, SupplieTrafficForm, BillingForm, DataSupplieForm, SearchFolio
from django.contrib.auth.decorators import login_required
from supplies.models import Supplies, SupplieStatus, WeekDays
from .filters import SupplieFilter
from .tables import SupplieTable
from django.http import HttpResponse
from openpyxl import Workbook
import io
import pytz
from datetime import datetime
from django.contrib.auth.models import User
from services.decorators import group_required
from django.utils import timezone
from django.forms import modelformset_factory
from django.http import JsonResponse
import json
from services.models import Truck,Trailer,Month

# Views
@login_required
@group_required('Administrative', 'SuperUsers')
def supplies(request):
    trips = Supplies.objects.filter(folio__isnull=True)
    return render(request, 'supplies.html', {
        'trips': trips
    })

@login_required
@group_required('Administrative', 'SuperUsers')
def supplies_finished(request):
    trips = Supplies.objects.filter(folio__isnull=False)
    return render(request, 'supplies_finished.html', {
        'trips': trips
    })

@login_required
@group_required('Administrative', 'SuperUsers')
def supplie_billing(request, id_serv):
    if request.method == "GET":
        trip = get_object_or_404(Supplies, pk=id_serv)
        form = BillingForm(instance=trip)
        return render(request,'supplie_billing.html', {
        'trip': trip,
        'form': form
        })
    else:
        try:
            trip = get_object_or_404(Supplies, pk=id_serv)
            form = BillingForm(request.POST, instance=trip)
            form.save()
            return redirect('supplies_finished')
        except:
            return render(request,'supplie_data.html', {
            'trip': trip,
            'form': form,
            'error': "No se pudo actualizar el registro",
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def supplie_data(request, id_serv):
    if request.method == "GET":
        trip = get_object_or_404(Supplies, pk=id_serv)
        form = DataSupplieForm(instance=trip)
        return render(request,'supplie_data.html', {
        'trip': trip,
        'form': form
        })
    else:
        try:
            trip = get_object_or_404(Supplies, pk=id_serv)
            form = DataSupplieForm(request.POST, instance=trip)
            form.save()
            return redirect('supplies')
        except:
            return render(request,'supplie_data.html', {
            'trip': trip,
            'form': form,
            'error': "No se pudo actualizar el registro",
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def supplie_traffic(request, id_serv):
    if request.method == "GET":
        trip = get_object_or_404(Supplies, pk=id_serv)
        form = SupplieTrafficForm(instance=trip)
        return render(request,'supplie_traffic.html', {
        'trip': trip,
        'form': form
        })
    else:
        trip = get_object_or_404(Supplies, pk=id_serv)
        form = SupplieTrafficForm(request.POST, instance=trip)
        form.save()
        return redirect('supplies')

@login_required
@group_required('Administrative', 'SuperUsers')
def supplie_editinit(request, id_serv):
    if request.method == "GET":
        trip = get_object_or_404(Supplies, pk=id_serv)
        form = CreateSupplieForm(instance=trip)
        return render(request,'supplie_editinit.html', {
        'trip': trip,
        'form': form
        })
    else:
        try:
            trip = get_object_or_404(Supplies, pk=id_serv)
            form = CreateSupplieForm(request.POST, instance=trip)
            form.save()
            return redirect('suplie_editinit')
        except:
            return render(request,'supplie_editinit.html', {
            'trip': trip,
            'form': form,
            'error': "No se pudo actualizar el registro",
            })

@login_required
@group_required('Administrative', 'SuperUsers')
def supplie_create(request):
    form = CreateSupplieForm
    if request.method == 'GET':
        return render(request, 'supplie_create.html',{
            'form': form
        })
    else:
        try:
            form = CreateSupplieForm(request.POST)
            new_trip = form.save(commit=False)
            new_trip.save()
            return redirect('supplie_create')
        except ValueError:
            return render(request, 'trip_create.html',{
            'form': form,
            'error': 'Por favor ingresa datos Validos'
        })

@group_required('Administrative', 'SuperUsers')
def supplie_delete(request, id_serv):
    trip = get_object_or_404(Supplies, pk=id_serv)
    if request.method == "POST":
        trip.delete()
        return redirect ('supplies')

@login_required
def status_supplie(request, id_serv):
    data = get_object_or_404(Supplies, pk=id_serv)

    if request.method == 'POST':
        form1 = StatusForm(request.POST)

        if form1.is_valid():
            status = form1.save(commit=False)
            status.service = data
            status.unit = data.truck
            status.driver = data.driver
            status.turn = data.turns
            status.supplier = data.port_supplier
            latitud = request.POST.get('latitud')
            longitud = request.POST.get('longitud')
            coordenadas = f"{latitud},{longitud}"
            status.coordinates = coordenadas
            status.save()
            return redirect('supplies_drivers')
        else:
            print(form1.errors)
            
    else:
        form1 = StatusForm()
        form2 = DataSupplieForm()

    return render(request, 'status_supplie.html', {'data': data, 'form1': form1, 'form2': form2})

def guardar_folio(request, id_serv):
    data = get_object_or_404(Supplies, pk=id_serv)  # Asegúrate de importar get_object_or_404
    if request.method == 'POST':
        folio = request.POST.get('folio')
        data.folio = folio
        data.save()
        return JsonResponse({'message': 'Folio guardado correctamente'}, status=200)
    return redirect('supplies_drivers.html',{
        'data': data
    })

@login_required
@group_required('Administrative', 'SuperUsers', 'Supplies')
def status_table(request):
    registros = SupplieStatus.objects.all().order_by('time_status')  # Ordena por hora de registro ascendente
    
    for i, registro in enumerate(registros):
        if i == len(registros) - 1:  # Último registro
            current_time = timezone.now()
            time_difference_last = current_time - registro.time_status
            registro.time_consult = current_time
            registro.time_dif = time_difference_last
            registro.save()
        elif i == len(registros) - 2:  # Penúltimo registro
            next_registro = registros[i + 1]
            registro.time_consult = next_registro.time_status
            time_difference = next_registro.time_status - registro.time_status
            registro.time_dif = time_difference
            registro.save()
        else:
            next_registro = registros[i + 1]
            time_difference = next_registro.time_status - registro.time_status
            registro.time_dif = time_difference
            registro.save()
    
    registro_filter = SupplieFilter(request.GET, queryset=registros)
    registro_table = SupplieTable(registro_filter.qs)
    context = {
        'registro_table': registro_table,
        'registro_filter': registro_filter,
    }
    
    return render(request, 'status_table.html', context)

@group_required('Administrative', 'SuperUsers', 'Supplies')
def supplie_excel(request):
    data = Supplies.objects.all()
    wb = Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="insumos.xlsx"'
    
    field_labels = {
        'folio': 'Folio',
        'week': 'Semana',
        'day_assignment': 'Dia de Asignacion',
        'driver': 'Operador',
        'truck': 'Unidad',
        'truck_plate': 'Placa de unidad',
        'trailer': 'Tolva',
        'trailer_plate': 'Placa de tolva',
        'capacity': 'Capacidad',
        'type_supplie': 'Tipo de servicio',
        'month': 'Mes',
        'serv_date': 'Fecha de servicio',
        'port_supplier': 'Proveedor',
        'turns': 'Turno',
    }
    
    for col_num, (field_name, label) in enumerate(field_labels.items(), start=1):
        ws.cell(row=1, column=col_num, value=label)
    
    for row_num, trip in enumerate(data, start=2):
        for col_num, field_name in enumerate(field_labels.keys(), start=1):
            field_value = getattr(trip, field_name)
            if field_name == 'day_assignment':  
                field_value = trip.day_assignment.name_day
            elif field_name == 'month':  
                field_value = trip.month.name_month
            elif field_name in ['time_status', 'status', 'unit', 'driver', 'turn', 'supplier','port_supplier' , 'extra_info', 'turns']:
                field_value = str(field_value)
            elif field_name == 'time_status':
                field_value = field_value.strftime('%Y-%m-%d %H:%M:%S')
            elif field_name == 'type_supplie':
                field_value = field_value.name_supplie  # Suponiendo que 'name' es la propiedad que contiene el nombre del tipo de suministro
            elif field_name in ['truck', 'truck_plate', 'trailer', 'trailer_plate']:
                field_value = getattr(getattr(trip, field_name), 'truck_num', '')  # Obtiene el número de camión
            ws.cell(row=row_num, column=col_num, value=field_value)
    
    output = io.BytesIO()
    wb.save(output)
    response.write(output.getvalue())
    return response

def status_excel(request):
    data = SupplieStatus.objects.all()
    wb = Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="status.xlsx"'
    
    field_labels = {
        'time_status': 'Folio',
        'time_consult': 'Semana',
        'time_dif': 'Dia de Asignacion',
        'status': 'Operador',
        'kms': 'Unidad',
        'coordinates': 'Placa de unidad',
        'unit': 'Tolva',
        'driver': 'Placa de tolva',
        'turn': 'Capacidad',
        'supplier': 'Tipo de servicio',
        'condition': 'Mes',
    }
    
    for col_num, (field_name, label) in enumerate(field_labels.items(), start=1):
        ws.cell(row=1, column=col_num, value=label)
    
    for row_num, status in enumerate(data, start=2):
        for col_num, field_name in enumerate(field_labels.keys(), start=1):
            if field_name == 'status':
                field_value = status.status.status
            else:
                field_value = getattr(status, field_name)
                if field_name in ['time_status', 'time_consult']:
                    field_value = field_value.strftime('%Y-%m-%d %H:%M:%S')
                elif field_name in ['unit', 'driver', 'turn', 'supplier']:
                    field_value = str(field_value)
                elif field_name == 'condition':
                    if field_value is not None:  # Verifica si condition no es None
                        field_value = field_value.condition_status
                    else:
                        field_value = ""  # Si condition es None, asigna un valor vacío
            ws.cell(row=row_num, column=col_num, value=field_value)

    
    output = io.BytesIO()
    wb.save(output)
    response.write(output.getvalue())
    return response    
    
@login_required
@group_required('Administrative', 'SuperUsers')
def create_assignment(request):
    if request.method == 'POST':
        form = CreateSupplieForm(request.POST)
        if form.is_valid():
            form.save()  
            return redirect('create_assignment')  
    else:
        form = CreateSupplieForm()

    return render(request, 'supplie_assignment.html', {'form': form})

@login_required
@group_required('Administrative', 'SuperUsers', 'Supplies')
def assignment_table(request):
    week_days = WeekDays.objects.all()
    week_number = request.GET.get('week_number', None)

    asignaciones_por_dia = {}

    if week_number is not None:
        for day in week_days:
            asignaciones = Supplies.objects.filter(week=week_number, day_assignment=day)
            asignaciones_por_dia[day] = asignaciones

    return render(request, 'table_assignment.html', {'week_days': week_days, 'asignaciones_por_dia': asignaciones_por_dia})

@login_required
def supplies_drivers(request):
    current_user = request.user.username
    trips = Supplies.objects.filter(
        truck__truck_num=current_user
    )

    return render(request, 'supplies_drivers.html', {
        'trips': trips
    })

@login_required
def search_folio(request):
    form = SearchFolio(request.GET or None)
    matches = []
    status_entries = []
    id_serv = None  # Inicializar con un valor predeterminado

    if form.is_valid():
        numero_folio = form.cleaned_data['folio']
        supplies = Supplies.objects.filter(folio=numero_folio)

        if supplies.exists():
            matches = supplies
            id_serv = supplies[0].id_serv

            # Buscar en el modelo SupplieStatus
            status_entries = SupplieStatus.objects.filter(service_id=id_serv)

    return render(request, 'search_folio.html', {'form': form, 'matches': matches, 'id_serv': id_serv, 'status_entries': status_entries})

